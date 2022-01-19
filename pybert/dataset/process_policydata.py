from openpyxl import load_workbook


j = 0
j=j+1

workbook = load_workbook("policydata.xlsx")
booksheet = workbook.active
rows = booksheet.rows
columns = booksheet.columns

i = 1
features = set()
with open('policy.txt', 'a', encoding='utf-8') as f:
    for row in rows:
        i= i+1
        cell_data_1 = booksheet.cell(row=i, column=2).value
        cell_data_2 = booksheet.cell(row=i, column=3).value
        if cell_data_1 is None or cell_data_2 is None or cell_data_1.isspace() or cell_data_2.isspace():
            continue
        features.add(cell_data_2.strip())
        f.write(cell_data_2.strip() +'\t'+cell_data_1.strip())
        f.write('\n')

# with open('policy_feature.txt', 'a', encoding='utf-8') as f:
#     for feature in features:
#         f.write(feature)
#         f.write('\n')



